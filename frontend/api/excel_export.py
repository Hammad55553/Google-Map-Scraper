import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import io

def generate_excel_from_leads(leads):
    # Convert leads to DataFrame
    data = []
    for lead in leads:
        data.append({
            "Business Name": lead.business_name,
            "Category": lead.category,
            "City": lead.city,
            "Rating": lead.rating,
            "Reviews Count": lead.reviews_count,
            "Phone": f"'{lead.phone}" if lead.phone else "",
            "Email": lead.email,
            "WhatsApp Link": str(lead.whatsapp_link) if lead.whatsapp_link else "",
            "Website": lead.website,
            "Address": lead.address,
            "Map URL": lead.map_url,
            "Lead Score": lead.lead_score,
            "Lead Grade": lead.lead_grade,
            "Status": lead.status,
            "Recommended Pitch": lead.recommended_pitch
        })
    df = pd.DataFrame(data)

    output = io.BytesIO()
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Leads')
        workbook = writer.book
        worksheet = writer.sheets['Leads']

        # Formatting
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_aligned_text = Alignment(horizontal="center", vertical="center")

        # Format Headers
        for col_num, value in enumerate(df.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned_text

        # Freeze Top Row
        worksheet.freeze_panes = 'A2'

        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            max_len = max((
                df[col].astype(str).map(len).max(),
                len(str(col))
            ))
            # Set a reasonable max width so columns don't get insanely wide
            adjusted_width = min(max_len + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = adjusted_width

    output.seek(0)
    return output
